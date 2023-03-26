import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook

ws = load_workbook("files/x3.xlsx")["Sheet1"]
max_row = ws.max_row

data = []
for i in range(1, max_row+1):
    data.append([ws[f"A{i}"].value, ws[f"B{i}"].value])

data = np.array(data)
x = data[:, 0]
y = data[:, 1]

dy = y[1:] - y[:-1]
dx = x[1:] - x[:-1]

yprime = y.copy()
yprime[1:] = dy/dx
yprime[0] = yprime[1]

wb_save = Workbook()
ws_save = wb_save.active

i = 0
for x_i in x:
    ws_save[f"A{i+1}"] = x_i
    ws_save[f"B{i+1}"] = yprime[i]
    i += 1

wb_save.save("files/3x2.xlsx")

ws2 = load_workbook("files/3x2.xlsx")["Sheet"]
max_row2 = ws2.max_row

data2 = []
for i in range(1, max_row2+1):
    data2.append([ws2[f"A{i}"].value, ws2[f"B{i}"].value])

data2 = np.array(data2)
dx = data2[:, 0]
dy = data2[:, 1]

ddy = dy[1:] - dy[:-1]
ddx = dx[1:] - dx[:-1]

ydoubleprime = yprime.copy()
ydoubleprime[1:] = ddy/ddx
ydoubleprime[0] = ydoubleprime[1]

wb_save2 = Workbook()
ws_save2 = wb_save2.active

k = 0
for dx_i in dx:
    ws_save2[f"A{k+1}"] = dx_i
    ws_save2[f"B{k+1}"] = ydoubleprime[k]
    k += 1

wb_save2.save("files/6x.xlsx")


plt.plot(x, y, label="x^3")
plt.plot(x, yprime, label="3x^2")
plt.plot(x, ydoubleprime, label="6x")
plt.legend()
plt.title("x^3 , 3x^2 , 6x")
plt.xlabel("x")
plt.ylabel("y")
plt.show()